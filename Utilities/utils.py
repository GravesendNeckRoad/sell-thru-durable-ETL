import io
import logging
import random
import time
from typing import List, Union

import openpyxl as xl
import pandas as pd

from azure.identity import DefaultAzureCredential
from azure.storage.blob import BlobServiceClient

class Style:
    """
    Simple styling tool to format an .xlsx worksheet with the openpyxl library.

    :param ws: (xl.worksheet.worksheet.Worksheet) The worksheet of the open workbook you wish to format

    Example:
    import openpyxl as xl
    wb = xl.load_workbook(path)
    ws = wb.active
    styler = Style(ws)
    """
    def __init__(self, ws) -> None:
        self.ws = ws

    def apply_styles_to_cell(self, cell: str, bold: bool = True, highlighter: bool = True, color: str = None) -> None:
        """
        Applies style formatting to a desired cell in the openpyxl worksheet.

        :param cell: (str) The cell you wish to format. (Ex: 'C2')
        :param bold: (bool) If true, the cell will be made bold
        :param highlighter: (bool) If true, the cell will be highlighted
        :param color: (str) The color to highlight your column. Default = yellow
        """
        if not color:
            color = 'FFFF00'

        if highlighter:
            self.ws[cell].fill = xl.styles.PatternFill(start_color=color, end_color=color, fill_type='solid')
        if bold:
            self.ws[cell].font = xl.styles.Font(bold=True)

    def change_font_color(self, cell: str, color: str = None) -> None:
        """
        Changes font color of a cell.
        
        :param color (str): The 8 digit aRGB hex value color you wish to change to.
        
        """
        self.ws[cell].font = xl.styles.Font(color=color)
        
        
    def currency_formatter(self, column_list: Union[List[str], str], max_row=None, currency: bool = True) -> None:
        """
        Formats every cell in a specified range as currency, or simply as a number with thousands separator.

        :param column_list: (str) The column letters you want to highlight (ex: ['K', 'L'], or 'E')
        :param max_row: (int) Specify a row number you want it to stop at. Default formats the entire column.
        :param currency: (bool) if False, then it will simply return the column formatted with a thousands separator.
        """
        if isinstance(column_list, str):
            column_list = list(column_list)

        if max_row is None:
            max_row = self.ws.max_row

        for col in column_list:
            for row in range(2, max_row + 1):
                if currency:
                    self.ws[f"{col}{row}"].number_format = '$#,##0.00'
                elif not currency:
                    self.ws[f"{col}{row}"].number_format = '#,##0'

    def align_and_center(self, start_row: int = 1, padding: int = 5) -> None:
        """ Auto align and widen all columns of your worksheet.

        :param start_row: (int) If your headers are long and/or text-wrapped, use >=2 to exclude headers as a reference.
        :param padding: (int) Add or remove whitespace from the columns
        """
        for row in self.ws.iter_cols(min_row=start_row):
            column_letter = xl.utils.get_column_letter(row[0].column)
            max_length = 0
            # first, widen each cell
            for cell in row:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
                self.ws.column_dimensions[column_letter].width = max_length + padding
                # then, center align
                cell.alignment = xl.styles.Alignment(
                    horizontal='center',
                    vertical='center'
                )

    def create_table(self, table_name: str = 'Table1') -> None:
        """Formats an Excel array as a table, by identifying the first/last rows and columns of the worksheet."""
        last_column = xl.utils.get_column_letter(len(list(self.ws.columns)))
        last_row = self.ws.max_row
        table_range = f"A1:{last_column}{last_row}"

        create_table = xl.worksheet.table.Table(
            displayName=f"{table_name}",
            ref=table_range
        )

        table_design = xl.worksheet.table.TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True
        )

        create_table.tableStyleInfo = table_design
        self.ws.add_table(create_table)

    def data_bars(self, column: str, color: str = '5e9bdd') -> None:
        """
        Creates data bars for a value column based on its min/max.
        
        Parameters:
            column: (str) the column LETTER you wish to format with databars
            color: (str) the color of the databars (default=blue)
        """
        # define the length of the specified column
        column_range = f"{column}2:{column}{self.ws.max_row}"
        
        # need to find the max value of the column
        max_value = 0
        for cell in self.ws[column][1:]:
            if isinstance(cell.value, (int, float)):
                if cell.value > max_value:
                    max_value = cell.value

        # create min/max rule
        rule = xl.formatting.rule.DataBarRule(
            start_type='num', 
            start_value=1,
            end_type='num',
            end_value=max_value, 
            color=color
            )

        # apply
        self.ws.conditional_formatting.add(column_range, rule)


class Helpers:
    def __init__(self):
        pass

    @staticmethod
    def exponential_backoff(n, rate_of_growth=1.5, base_seconds=2, jitter=.01) -> None:
        """Simple timer function to manage API throttling."""
        x = (base_seconds * (rate_of_growth ** n))
        y = (random.uniform(-jitter*x, jitter*x))
        print(f"\tRetry attempt #{n} - {x+y:.2f} seconds ...")
        logging.info(f"\tRetry attempt #{n} - {x+y:.2f} seconds ...")
        time.sleep(x+y)

    @staticmethod
    def save_df_to_mem(df: pd.DataFrame) -> io.BytesIO:
        """Saves a Pandas DataFrame as .xlsx to an in-memory buffer"""
        buffer = io.BytesIO()
        df.to_excel(buffer, index=False)
        buffer.seek(0)
        return buffer


class BlobHandler:
    """
    Instantiates BlobServiceClient and writes data to/from a specified blob container
    
    Parameters:
        -storage_account: (str) Name of the storage account you wish to pull from
        -container_name: (str) Name of the blob container within the above specified storage account 
    
    Considerations:
        -This class uses DefaultAzureCredential(), so make sure your managed identities are in order
    """
    def __init__(self, storage_account: str, container_name: str):
        self.storage_account = storage_account
        self.container_name = container_name
        self.blob_service_client = self.__init_blob_client()
    
    def __init_blob_client(self) -> BlobServiceClient:
        """Private method: initiates and validates a blob client upon class instantiation. Returns client object"""        
        try:
            return BlobServiceClient(
                account_url=f"https://{self.storage_account}.blob.core.windows.net/", 
                credential=DefaultAzureCredential()
                )
        except Exception as e:
            logging.error(f"Could not validate the BlobServiceClient: {str(e)}")
            raise
            
    def save_to_blob(self, buffer: io.BytesIO, save_as: str) -> None:
        """Uploads in-memory buffer file to the blob container, titled after the save_as parameter"""
        
        if not isinstance(buffer, io.BytesIO):
            raise TypeError("The data passed to this method must be of io.BytesIO type")

        try:
            blob_client = self.blob_service_client.get_blob_client(
                container=self.container_name, 
                blob=save_as
                )
            blob_client.upload_blob(buffer, overwrite=True)
            logging.info(f"Uploaded file '{save_as}' to the designated blob container")

        except Exception as e:
            logging.error(f"Could not save {save_as} to blob. {str(e)}")
            raise
   
    def get_from_blob(self, blob_name: str) -> pd.DataFrame:
        """Transfer from blob to local machine, returns Pandas df. Blob name should include file extension"""
        try:
            blob_client = self.blob_service_client.get_blob_client(
                container=self.container_name, 
                blob=blob_name
                )
            blob_data = blob_client.download_blob().readall()

            if blob_name.endswith('xlsx'):
                df = pd.read_excel(io.BytesIO(blob_data), engine='openpyxl')
            elif blob_name.endswith('csv'):
                df = pd.read_csv(io.BytesIO(blob_data))
            elif blob_name.endswith('tsv'):
                df = pd.read_csv(io.BytesIO(blob_data), sep='\t')
            elif blob_name.endswith('txt'):
                txt_file = io.BytesIO(blob_data).read().decode('utf-8')
                df = pd.DataFrame(txt_file.splitlines())
            else:
                raise TypeError("Only supports xlsx/csv/tsv/txt files for now")

            return df
        
        except Exception as e:
            logging.error(f"Error getting your file from blob. {str(e)}")
            raise
